# カラー・スタイル定数
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── カラーパレット ──────────────────────────────
HEADER_BG      = "1A1A2E"
LUNCH_SECTION  = "1565C0"
DINNER_SECTION = "4A148C"

COL_PRICE_REG  = "C62828"
COL_DISCOUNT   = "AD1457"
COL_PRICE_PAY  = "1565C0"

ROW_DISCOUNT   = "FFF0F0"
ROW_ALLERGY    = "FFF8E1"
ROW_DINNER_A   = "F3E5F5"
ROW_DINNER_B   = "EDE7F6"
ROW_LUNCH_A    = "E3F2FD"
ROW_LUNCH_B    = "FAFEFF"

ALERT_TITLE    = "B71C1C"
ALERT_BG       = "FFF3F3"
WARN_BG        = "FFFDE7"

AT_SECTION     = "FFF0E6"
AT_COUNT       = "FFCCBC"

WHITE          = "FFFFFF"
BLACK          = "000000"

# ── ヘルパー関数 ───────────────────────────────
def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def font(color: str = BLACK, bold: bool = False, size: int = 10, name: str = "Arial") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def red_border() -> Border:
    s = Side(style="medium", color=COL_DISCOUNT)
    return Border(left=s, right=s, top=s, bottom=s)

# ── 列定義（新レイアウト）─────────────────────────
# A: 席  B: 時間  C: 名前  D: 人数  E: 媒体
# F: 割引・ポイント  G: プラン
# H: 正規金額(非表示)  I: 支払金額(非表示)
# J: アレルギー  K: 記念日メッセージ  L: 備考・要望  M: 目的
COLUMNS = [
    ("席",              6),   # A  idx=0
    ("時間",            7),   # B  idx=1
    ("名前",           16),   # C  idx=2
    ("人数",            5),   # D  idx=3
    ("媒体",           11),   # E  idx=4
    ("割引・ポイント",  15),   # F  idx=5  ← 媒体とプランの間
    ("プラン",         30),   # G  idx=6
    ("正規金額(税込)",  12),   # H  idx=7  ← 非表示
    ("支払金額(税込)",  12),   # I  idx=8  ← 非表示
    ("アレルギー",      20),   # J  idx=9
    ("記念日メッセージ", 28),  # K  idx=10
    ("備考・要望",      32),   # L  idx=11
    ("目的",            10),   # M  idx=12
]
COL_NAMES  = [c[0] for c in COLUMNS]
COL_WIDTHS = [c[1] for c in COLUMNS]

# 非表示列インデックス（0始まり）
HIDDEN_COL_IDXS = [7, 8]   # H=正規金額, I=支払金額

# 金額関連列インデックス（0始まり）
IDX_DISCOUNT   = COL_NAMES.index("割引・ポイント")   # F=5
IDX_PRICE_REG  = COL_NAMES.index("正規金額(税込)")   # H=7
IDX_PRICE_PAY  = COL_NAMES.index("支払金額(税込)")   # I=8
