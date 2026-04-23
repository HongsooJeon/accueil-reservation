"""
Sheet 1: 予約管理（メインシート）
新列レイアウト:
  A:席  B:時間  C:名前  D:人数  E:媒体  F:割引・ポイント  G:プラン
  H:正規金額[非表示]  I:支払金額[非表示]
  J:アレルギー  K:記念日メッセージ  L:備考・要望  M:目的
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .styles import (
    fill, font, center, left, thin_border, red_border,
    HEADER_BG, LUNCH_SECTION, DINNER_SECTION,
    COL_DISCOUNT, COL_PRICE_REG, COL_PRICE_PAY,
    ROW_DISCOUNT, ROW_ALLERGY, ROW_DINNER_A, ROW_DINNER_B, ROW_LUNCH_A, ROW_LUNCH_B,
    WHITE, BLACK,
    COLUMNS, COL_NAMES, COL_WIDTHS, HIDDEN_COL_IDXS,
    IDX_DISCOUNT, IDX_PRICE_REG, IDX_PRICE_PAY,
)
from parser.pdf_parser import Reservation

DOW_JA       = ["月", "火", "水", "木", "金", "土", "日"]
DINNER_START = "16:00"

# ヘッダー列名（改行あり）
_HEADER_NAMES = [
    "席", "時間", "名前", "人数", "媒体",
    "割引・\nポイント", "プラン",
    "正規金額\n(税込)", "支払金額\n(税込)",
    "アレルギー", "記念日メッセージ", "備考・要望", "目的",
]


def _get_dow(date_str: str) -> str:
    dt = datetime(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:]))
    return DOW_JA[dt.weekday()]


def _is_dinner(time_str: str) -> bool:
    return time_str >= DINNER_START


def _fmt_money(val):
    return val if val is not None else "—"


def _fmt_discount(val) -> str:
    if val is None or val == 0:
        return "—"
    return f"－{val:,}円"


def _anniv_cell(r: Reservation):
    """アレルギーは別列になったので、記念日メッセージのみ返す"""
    if r.source == "OZmall":
        return r.anniversary_msg if r.anniversary_msg else "（メッセージなし）"
    return r.anniversary_msg if r.anniversary_msg else None


def _allergy_cell(r: Reservation):
    """J列：アレルギー内容テキスト"""
    if r.has_allergy:
        return r.allergy_detail if r.allergy_detail else "要確認"
    return None


def _notes_cell(r: Reservation):
    if r.source == "ヒトサラ":
        return "⚠ 当日ポイント利用の可能性あり・要確認"
    return r.notes if r.notes else None


def _table_cell(r: Reservation) -> str:
    return r.table if r.table else "—"


def _row_color(r: Reservation, odd: bool) -> str:
    if r.discount:
        return ROW_DISCOUNT
    if r.has_allergy:
        return ROW_ALLERGY
    if _is_dinner(r.time_start):
        return ROW_DINNER_A if not odd else ROW_DINNER_B
    return ROW_LUNCH_A if not odd else ROW_LUNCH_B


def _calc_row_height(values: list) -> float:
    """wrap_text前提で行高を推定する。
    全角文字は半角2幅として計算し、列幅（半角文字単位）で割って行数を求める。
    最小18pt を保証する。
    """
    import math
    # 各列の幅（半角文字単位）: A〜M の順
    col_widths_approx = [6, 7, 16, 5, 11, 15, 30, 12, 12, 20, 28, 32, 10]
    max_lines = 1
    for ci, val in enumerate(values):
        if not val or not isinstance(val, str):
            continue
        w = col_widths_approx[ci] if ci < len(col_widths_approx) else 12
        # 全角=2幅、半角=1幅の合計を列幅で割り上げて行数を算出
        char_units = sum(2 if ord(c) > 0x7F else 1 for c in val)
        explicit_newlines = val.count('\n')
        lines = math.ceil(char_units / w) + explicit_newlines
        lines = max(1, lines)
        max_lines = max(max_lines, lines)
    # 1行≈15pt、最小18pt（ユーザー指定）、最大90pt
    return max(18, min(90, max_lines * 15))


def build_sheet_main(wb: Workbook, reservations: list[Reservation], date_str: str):
    ws = wb.active
    ws.title = "予約管理"

    n_cols = len(COLUMNS)
    dow  = _get_dow(date_str)
    yyyy, mm, dd = date_str[:4], date_str[4:6], date_str[6:]

    total_groups = len(reservations)
    total_people = sum(r.party_size for r in reservations)

    # ─ Row 1: タイトル ────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = f"📅  カフェ アクイーユ 恵比寿\u3000|\u3000予約管理表\u3000{yyyy}.{mm}.{dd}（{dow}）"
    c.fill      = fill(HEADER_BG)
    c.font      = font(WHITE, bold=True, size=14)
    c.alignment = center()
    ws.row_dimensions[1].height = 32

    # ─ Row 2: サマリー ───────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = (
        f"本日計：{total_groups}組 / {total_people}名\u3000\u3000"
        f"⚠ 赤枠行 = 会計時に割引・ポイント適用あり\u3000要注意"
    )
    c.fill      = fill("F5F5F5")
    c.font      = font(BLACK, size=10)
    c.alignment = left()
    ws.row_dimensions[2].height = 18

    # ─ Row 3: ヘッダー ───────────────────────────────────────
    price_color_map = {
        IDX_DISCOUNT:  COL_DISCOUNT,
        IDX_PRICE_REG: COL_PRICE_REG,
        IDX_PRICE_PAY: COL_PRICE_PAY,
    }
    for ci, name in enumerate(_HEADER_NAMES):
        c = ws.cell(row=3, column=ci + 1)
        c.value     = name
        c.fill      = fill(price_color_map.get(ci, HEADER_BG))
        c.font      = font(WHITE, bold=True, size=10)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[3].height = 28

    # 列幅 & 非表示設定
    for ci, width in enumerate(COL_WIDTHS):
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = width
        if ci in HIDDEN_COL_IDXS:
            ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A4"

    # ─ データ行 ──────────────────────────────────────────────
    lunch_rsvs  = [r for r in reservations if not _is_dinner(r.time_start)]
    dinner_rsvs = [r for r in reservations if _is_dinner(r.time_start)]

    data_row = 4
    lunch_sub_row  = None   # ランチ小計行の行番号
    dinner_sub_row = None   # ディナー小計行の行番号
    price_start    = 4      # SUMIF 用データ開始行

    def _write_data_rows(rsvs, section_color, section_label):
        nonlocal data_row
        if not rsvs:
            return None, None
        _write_section_bar(ws, section_label, data_row, section_color, n_cols)
        data_row += 1
        start = data_row
        for idx, r in enumerate(rsvs):
            bg = _row_color(r, idx % 2 == 1)
            values = [
                _table_cell(r),
                r.time_start,
                r.name if r.name else None,
                r.party_size if r.party_size else None,
                r.source if r.source else None,
                _fmt_discount(r.discount),
                r.plan if r.plan else None,
                _fmt_money(r.price_regular),
                _fmt_money(r.price_pay),
                _allergy_cell(r),
                _anniv_cell(r),
                _notes_cell(r),
                r.purpose if r.purpose else None,
            ]
            center_cols = {0, 1, 3}
            for ci, val in enumerate(values):
                c = ws.cell(row=data_row, column=ci + 1)
                c.value     = val
                c.fill      = fill(bg)
                c.border    = thin_border()
                c.alignment = center() if ci in center_cols else left()
                if ci == 9 and val:
                    c.font = font(COL_PRICE_REG, bold=True, size=10)
                else:
                    c.font = font(BLACK, size=10)
                if r.discount and ci == IDX_DISCOUNT:
                    c.border = red_border()
            ws.row_dimensions[data_row].height = _calc_row_height(values)
            data_row += 1
        return start, data_row - 1   # (data_start, data_end)

    # ── ランチ ──
    l_start, l_end = _write_data_rows(
        lunch_rsvs, LUNCH_SECTION, "▶  LUNCH  11:00 – 15:30"
    )
    if l_start:
        _write_subtotal_bar(ws, "ランチ 小計", data_row, l_start, l_end, LUNCH_SECTION, n_cols)
        lunch_sub_row = data_row
        data_row += 1

    # ── ディナー ──
    d_start, d_end = _write_data_rows(
        dinner_rsvs, DINNER_SECTION, "▶  DINNER  16:00 – 22:00"
    )
    if d_start:
        _write_subtotal_bar(ws, "ディナー 小計", data_row, d_start, d_end, DINNER_SECTION, n_cols)
        dinner_sub_row = data_row
        data_row += 1

    # ─ 1日合計行 ──────────────────────────────────────────────
    last_data = data_row - 1

    # 件数: ランチ+ディナー の小計セルC列を参照
    if lunch_sub_row and dinner_sub_row:
        cnt_formula  = f"=C{lunch_sub_row}+C{dinner_sub_row}"
        ppl_formula  = f"=D{lunch_sub_row}+D{dinner_sub_row}"
    elif lunch_sub_row:
        cnt_formula  = f"=C{lunch_sub_row}"
        ppl_formula  = f"=D{lunch_sub_row}"
    elif dinner_sub_row:
        cnt_formula  = f"=C{dinner_sub_row}"
        ppl_formula  = f"=D{dinner_sub_row}"
    else:
        cnt_formula  = 0
        ppl_formula  = 0

    totals = [
        "1日 合計",
        None,
        cnt_formula,                                         # C: 件数
        ppl_formula,                                         # D: 人数
        None, None, None,
        f"=SUMIF(H{price_start}:H{last_data},\">0\")",      # H: 正規金額合計
        f"=SUMIF(I{price_start}:I{last_data},\">0\")",      # I: 支払金額合計
        None, None, None, None,
    ]
    for ci, val in enumerate(totals):
        c = ws.cell(row=data_row, column=ci + 1)
        c.value     = val
        c.fill      = fill(HEADER_BG)
        c.font      = font(WHITE, bold=True, size=10)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[data_row].height = 22

    # 印刷設定：A4縦・1枚に収まるように自動縮小
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4        # A4
    ws.page_setup.orientation = "portrait"             # 縦
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1                      # 横1ページ
    ws.page_setup.fitToHeight = 1                      # 縦1ページ
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _write_subtotal_bar(ws: Worksheet, label: str, row: int,
                        data_start: int, data_end: int,
                        color: str, n_cols: int):
    """ランチ・ディナー小計行を書く。C列=件数(COUNTA)、D列=人数(SUM)。"""
    # A-B 結合: ラベル
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value     = label
    c.fill      = fill(color)
    c.font      = font(WHITE, bold=True, size=10)
    c.alignment = center()
    c.border    = thin_border()

    # C: 件数（名前列=C をカウント）
    c = ws.cell(row=row, column=3)
    c.value     = f"=COUNTA(C{data_start}:C{data_end})"
    c.fill      = fill(color)
    c.font      = font(WHITE, bold=True, size=10)
    c.alignment = center()
    c.border    = thin_border()

    # D: 人数（D列を合計）
    c = ws.cell(row=row, column=4)
    c.value     = f"=SUM(D{data_start}:D{data_end})"
    c.fill      = fill(color)
    c.font      = font(WHITE, bold=True, size=10)
    c.alignment = center()
    c.border    = thin_border()

    # E-M: 空セル（スタイルのみ）
    for col in range(5, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = fill(color)
        c.border = thin_border()

    ws.row_dimensions[row].height = 20


def _write_section_bar(ws: Worksheet, label: str, row: int, color: str, n_cols: int):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws[f"A{row}"]
    c.value     = label
    c.fill      = fill(color)
    c.font      = font(WHITE, bold=True, size=11)
    c.alignment = left()
    ws.row_dimensions[row].height = 20
