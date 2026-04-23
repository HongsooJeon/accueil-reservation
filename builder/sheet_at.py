"""
Sheet 3: AT集計（製造用）
"""
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .styles import (
    fill, font, center, left, thin_border,
    HEADER_BG, AT_SECTION, AT_COUNT, WHITE, BLACK,
)
from parser.pdf_parser import Reservation, is_at_plan, is_aniva_plan

_HEADERS = [
    ("時間帯",             8),
    ("ATセット\n台数(人)", 12),
    ("アニバ\nランチ",     10),
    ("合計\n来客",         10),
    ("予約者 / プラン 詳細", 42),
    ("備考",               12),
]


def _time_slot(time_str: str) -> str:
    try:
        h = int(time_str.split(":")[0])
        return f"{h:02d}:00"
    except Exception:
        return time_str


def _plan_short(plan: str) -> str:
    return re.sub(r'\((?:平日|週末|土日|期間限定)[^)]*\)$', '', plan).strip()


def build_sheet_at(wb: Workbook, reservations: list[Reservation], date_str: str):
    ws = wb.create_sheet("AT集計（製造用）")
    n = len(_HEADERS)
    yyyy, mm, dd = date_str[:4], date_str[4:6], date_str[6:]

    # ─ Row 1: タイトル ───────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    c = ws["A1"]
    c.value     = f"🍓  いちごみるく AT  ／  時間帯別 製造数一覧\u3000{yyyy}.{mm}.{dd}"
    c.fill      = fill(AT_SECTION)
    c.font      = font(BLACK, bold=True, size=13)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # ─ Row 2: サブタイトル ──────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    c = ws["A2"]
    c.value = (
        "※ ATセット＝いちごみるくアフタヌーンティー本体の台数（人数ベース）"
        "\u3000アニバーサリーランチは別枠"
    )
    c.fill      = fill("FFFAF0")
    c.font      = font(BLACK, bold=False, size=9)
    c.alignment = left()
    ws.row_dimensions[2].height = 16

    # ─ Row 3: ヘッダー ───────────────────────────────────────
    for ci, (name, width) in enumerate(_HEADERS):
        c = ws.cell(row=3, column=ci + 1)
        c.value     = name
        c.fill      = fill(HEADER_BG)
        c.font      = font(WHITE, bold=True, size=10)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci + 1)].width = width
    ws.row_dimensions[3].height = 28

    ws.freeze_panes = "A4"

    # ─ 集計 ──────────────────────────────────────────────────
    slot_data: dict[str, dict] = defaultdict(lambda: {
        "at_people": 0,
        "aniva_people": 0,
        "total": 0,
        "details": [],
    })

    at_reservations = [r for r in reservations if is_at_plan(r.plan) or is_aniva_plan(r.plan)]

    for r in at_reservations:
        slot = _time_slot(r.time_start)
        d = slot_data[slot]
        d["total"] += r.party_size

        if is_aniva_plan(r.plan):
            d["aniva_people"] += r.party_size
        else:
            d["at_people"] += r.party_size
            short = _plan_short(r.plan)
            d["details"].append(f"{r.time_start}  {r.name}（{r.party_size}名）  ▸ {short}")

    # ─ データ行 ──────────────────────────────────────────────
    data_row = 4
    for slot in sorted(slot_data.keys()):
        d = slot_data[slot]
        details_text = "\n".join(d["details"]) if d["details"] else "—"
        aniva_val    = d["aniva_people"] if d["aniva_people"] > 0 else "—"
        row_height   = max(18, 15 * max(len(d["details"]), 1))

        values = [
            slot,
            d["at_people"],
            aniva_val,
            d["total"],
            details_text,
            None,
        ]

        for ci, val in enumerate(values):
            c = ws.cell(row=data_row, column=ci + 1)
            c.value  = val
            c.fill   = fill(AT_COUNT) if ci == 1 else fill("FFFFFF")
            c.border = thin_border()
            c.font   = font(BLACK, size=10)
            if ci == 4:
                c.alignment = left()
            else:
                c.alignment = center()

        ws.row_dimensions[data_row].height = row_height
        data_row += 1

    # ─ 合計行 ────────────────────────────────────────────────
    total_at    = sum(d["at_people"]    for d in slot_data.values())
    total_aniva = sum(d["aniva_people"] for d in slot_data.values())
    total_all   = sum(d["total"]        for d in slot_data.values())

    totals = ["合 計", total_at, total_aniva or "—", total_all, None, None]
    for ci, val in enumerate(totals):
        c = ws.cell(row=data_row, column=ci + 1)
        c.value     = val
        c.fill      = fill(HEADER_BG)
        c.font      = font(WHITE, bold=True, size=10)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[data_row].height = 20

    if not at_reservations:
        ws.merge_cells(f"A4:{get_column_letter(n)}4")
        c = ws["A4"]
        c.value     = "本日の対象予約なし"
        c.font      = font(BLACK, size=10)
        c.alignment = center()
