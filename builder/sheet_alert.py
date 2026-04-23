"""
Sheet 2: ⚠ 会計注意リスト
"""
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .styles import (
    fill, font, center, left, thin_border,
    HEADER_BG, ALERT_TITLE, ALERT_BG, WARN_BG, WHITE, BLACK,
    COL_PRICE_REG, COL_DISCOUNT, COL_PRICE_PAY,
)
from parser.pdf_parser import Reservation

ALERT_COLS = [
    ("来店\n時間",          6),
    ("退席\n目安",          6),
    ("お名前",             14),
    ("人\n数",              5),
    ("席",                  6),
    ("媒体",                8),
    ("プラン名",            22),
    ("正規金額\n(税込)",    10),
    ("割引・ポイント\n種別", 16),
    ("支払金額\n(税込)",    10),
    ("スタッフ備考・対応メモ", 30),
]
ALERT_COL_NAMES  = [c[0] for c in ALERT_COLS]
ALERT_COL_WIDTHS = [c[1] for c in ALERT_COLS]

_PRICE_COL_MAP = {
    7:  COL_PRICE_REG,
    8:  COL_DISCOUNT,
    9:  COL_PRICE_PAY,
}

_FOOTER_LINES = [
    "【ぐるなび / ヒトサラ ご利用のお客様へ】"
    "当日来店時に「ポイントをご利用されますか？」と必ず口頭確認してください。"
    "予約票に記載がない場合でも適用される場合があります。",
    "【OZmallご利用のお客様へ】"
    "OZポイント・OZの日割引は事前に金額が確定しています。"
    "「支払金額」列の金額で会計してください。",
]


def _calc_row_height(row: list) -> float:
    """wrap_text前提で行高を推定（全角=2幅、半角=1幅）。最小18pt保証。"""
    # ALERT_COLS の列幅（半角文字単位）
    col_widths = [6, 6, 14, 5, 6, 8, 22, 10, 16, 10, 30]
    max_lines = 1
    for ci, val in enumerate(row):
        if not val or not isinstance(val, str):
            continue
        w = col_widths[ci] if ci < len(col_widths) else 10
        char_units = sum(2 if ord(c) > 0x7F else 1 for c in val)
        lines = max(1, math.ceil(char_units / w) + val.count('\n'))
        max_lines = max(max_lines, lines)
    return max(18, min(90, max_lines * 15))


def _is_alert(r: Reservation) -> bool:
    if r.discount and r.discount > 0:
        return True
    if r.source == "ヒトサラ":
        return True
    return False


def _build_row(r: Reservation) -> list:
    if r.source == "ヒトサラ":
        return [
            r.time_start,
            r.time_end if r.time_end else "",
            r.name,
            r.party_size if r.party_size else "",
            r.table if r.table else "—",
            r.source,
            r.plan if r.plan else "",
            "当日精算",
            "⚠ 当日ポイント使用の可能性あり・来店時に確認",
            "要確認",
            "来店時にポイント利用の有無を口頭確認すること",
        ]
    # OZmall 割引あり
    return [
        r.time_start,
        r.time_end if r.time_end else "",
        r.name,
        r.party_size if r.party_size else "",
        r.table if r.table else "—",
        r.source,
        r.plan if r.plan else "",
        r.price_regular if r.price_regular is not None else "—",
        r.discount_label if r.discount_label else f"割引　－{r.discount:,}円",
        r.price_pay if r.price_pay is not None else "—",
        "会計時に割引を適用すること",
    ]


def build_sheet_alert(wb: Workbook, reservations: list[Reservation], date_str: str):
    ws = wb.create_sheet("⚠ 会計注意リスト")
    n = len(ALERT_COLS)
    yyyy, mm, dd = date_str[:4], date_str[4:6], date_str[6:]

    # ─ Row 1: タイトル ───────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    c = ws["A1"]
    c.value     = f"⚠   会計注意リスト\u3000｜\u3000割引・ポイント適用あり\u3000{yyyy}.{mm}.{dd}"
    c.fill      = fill(ALERT_TITLE)
    c.font      = font(WHITE, bold=True, size=13)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # ─ Row 2: 説明文 ────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    c = ws["A2"]
    c.value = (
        "このシートに記載のお客様は、会計時に必ず割引・ポイントを適用してください。"
        "見落としは再来店しない理由になります。"
    )
    c.fill      = fill("FFF3F3")
    c.font      = font(ALERT_TITLE, bold=False, size=10)
    c.alignment = left()
    ws.row_dimensions[2].height = 18

    # ─ Row 3: ヘッダー ───────────────────────────────────────
    for ci, name in enumerate(ALERT_COL_NAMES):
        c = ws.cell(row=3, column=ci + 1)
        c.value     = name
        c.fill      = fill(_PRICE_COL_MAP.get(ci, HEADER_BG))
        c.font      = font(WHITE, bold=True, size=10)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[3].height = 28

    # 列幅
    for ci, width in enumerate(ALERT_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(ci + 1)].width = width

    ws.freeze_panes = "A4"

    # ─ データ行 ──────────────────────────────────────────────
    data_row = 4
    center_cols = {0, 1, 3}

    for r in reservations:
        if not _is_alert(r):
            continue

        bg  = WARN_BG if r.source == "ヒトサラ" else ALERT_BG
        row = _build_row(r)

        for ci, val in enumerate(row):
            c = ws.cell(row=data_row, column=ci + 1)
            c.value     = val
            c.fill      = fill(bg)
            c.border    = thin_border()
            c.alignment = center() if ci in center_cols else left()
            c.font      = font(BLACK, size=10)

        ws.row_dimensions[data_row].height = _calc_row_height(row)
        data_row += 1

    if data_row == 4:
        ws.merge_cells(f"A4:{get_column_letter(n)}4")
        c = ws["A4"]
        c.value     = "対象なし"
        c.font      = font(BLACK, size=10)
        c.alignment = center()
        data_row = 5

    # ─ フッター ──────────────────────────────────────────────
    data_row += 1
    for line in _FOOTER_LINES:
        ws.merge_cells(f"A{data_row}:{get_column_letter(n)}{data_row}")
        c = ws[f"A{data_row}"]
        c.value     = line
        c.font      = font(ALERT_TITLE if line.startswith("【") else BLACK, bold=line.startswith("【"), size=10)
        c.alignment = left()
        ws.row_dimensions[data_row].height = 18
        data_row += 1
